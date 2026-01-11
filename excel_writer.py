"""
Excel Writer Module
Fills an Excel template with extracted invoice data using configurable mapping.
"""

import io
from copy import copy
from typing import Optional

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, Alignment

from config import HEADER_MAPPING, ITEMS_MAPPING, ITEMS_START_ROW


def create_template_workbook() -> Workbook:
    """Create a new workbook with the standard template headers."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = [
        ("A1", "Số hóa đơn"),
        ("B1", "Ngày hóa đơn"),
        ("C1", "MST Bên bán"),
        ("D1", "Tên người bán"),
        ("E1", "MST bên mua"),
        ("F1", "Tên người mua"),
        ("G1", "Tên hàng hóa"),
        ("H1", "Đơn vị"),
        ("I1", "Số lượng"),
        ("J1", "Đơn giá"),
        ("K1", "% VAT"),
        ("L1", "Thành tiền trước VAT"),
        ("M1", "VAT"),
        ("N1", "Thành tiền sau VAT"),
    ]

    # Set headers with formatting
    for cell_ref, value in headers:
        cell = ws[cell_ref]
        cell.value = value
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Set column widths
    column_widths = {
        "A": 12, "B": 14, "C": 15, "D": 35, "E": 15, "F": 35,
        "G": 30, "H": 10, "I": 10, "J": 15, "K": 8, "L": 18, "M": 15, "N": 18
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    return wb


def fill_excel_template(
    invoice_data: dict,
    template_bytes: Optional[bytes] = None
) -> bytes:
    """
    Fill Excel template with invoice data.

    Args:
        invoice_data: Extracted and normalized invoice data
        template_bytes: Optional existing template file bytes. If None, creates new template.

    Returns:
        Bytes of the filled Excel file
    """
    # Load existing template or create new one
    if template_bytes:
        wb = load_workbook(io.BytesIO(template_bytes))
        ws = wb.active
    else:
        wb = create_template_workbook()
        ws = wb.active

    items = invoice_data.get("items", [])

    # Fill data for each item (one row per item)
    for idx, item in enumerate(items):
        row_num = ITEMS_START_ROW + idx

        # Fill header fields (repeated for each row in this format)
        for json_field, excel_col in HEADER_MAPPING.items():
            cell = ws[f"{excel_col}{row_num}"]
            value = invoice_data.get(json_field)
            if value is not None:
                cell.value = value

        # Fill item-specific fields
        for json_field, excel_col in ITEMS_MAPPING.items():
            cell = ws[f"{excel_col}{row_num}"]
            value = item.get(json_field)
            if value is not None:
                cell.value = value

                # Format numeric cells
                if json_field in ["quantity", "unit_price", "amount_before_vat", "vat_amount", "amount_after_vat"]:
                    cell.number_format = "#,##0"
                elif json_field == "vat_rate":
                    cell.number_format = "0%"
                    # Convert percentage to decimal for Excel if it's a whole number
                    if isinstance(value, (int, float)) and value > 1:
                        cell.value = value / 100

    # If no items, still fill header info in row 2
    if not items:
        row_num = ITEMS_START_ROW
        for json_field, excel_col in HEADER_MAPPING.items():
            cell = ws[f"{excel_col}{row_num}"]
            value = invoice_data.get(json_field)
            if value is not None:
                cell.value = value

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def format_invoice_number(series: Optional[str], number: Optional[str]) -> str:
    """Format invoice number with series prefix if available."""
    if series and number:
        return f"{series}-{number}"
    return number or ""


def fill_excel_batch(
    invoices_data: list[dict],
    template_bytes: Optional[bytes] = None
) -> bytes:
    """
    Fill Excel template with multiple invoices data.
    All invoices are combined into a single sheet.

    Args:
        invoices_data: List of extracted and normalized invoice data dicts
        template_bytes: Optional existing template file bytes. If None, creates new template.

    Returns:
        Bytes of the filled Excel file
    """
    # Load existing template or create new one
    if template_bytes:
        wb = load_workbook(io.BytesIO(template_bytes))
        ws = wb.active
    else:
        wb = create_template_workbook()
        ws = wb.active

    current_row = ITEMS_START_ROW

    for invoice_data in invoices_data:
        items = invoice_data.get("items", [])

        if items:
            # Fill data for each item (one row per item)
            for item in items:
                # Fill header fields (repeated for each row in this format)
                for json_field, excel_col in HEADER_MAPPING.items():
                    cell = ws[f"{excel_col}{current_row}"]
                    value = invoice_data.get(json_field)
                    if value is not None:
                        cell.value = value

                # Fill item-specific fields
                for json_field, excel_col in ITEMS_MAPPING.items():
                    cell = ws[f"{excel_col}{current_row}"]
                    value = item.get(json_field)
                    if value is not None:
                        cell.value = value

                        # Format numeric cells
                        if json_field in ["quantity", "unit_price", "amount_before_vat", "vat_amount", "amount_after_vat"]:
                            cell.number_format = "#,##0"
                        elif json_field == "vat_rate":
                            cell.number_format = "0%"
                            # Convert percentage to decimal for Excel if it's a whole number
                            if isinstance(value, (int, float)) and value > 1:
                                cell.value = value / 100

                current_row += 1
        else:
            # No items, still fill header info in one row
            for json_field, excel_col in HEADER_MAPPING.items():
                cell = ws[f"{excel_col}{current_row}"]
                value = invoice_data.get(json_field)
                if value is not None:
                    cell.value = value
            current_row += 1

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
